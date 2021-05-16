# -*- coding: utf-8 -*-
"""Covid-19 deaths internal component.

Module containing operations with Covid-19 deaths.

Example:
    Italian covid deaths can be fetched with
    
        covid_deaths_it = deaths.covid19italy.covid_deaths()

    Get Covid-19 deaths for all countries with
    
        covid_deaths = deaths.get_data()
    
    Construct a violin plot of Covid-19 deaths with
    
        deaths.plot_violin()
    
    Test that age of dying on Covid-19 is greater than 60 with
    
        result_over60 = deaths.test_over60()
    
"""
import covid19poland
import covid19czechia
import io
import matplotlib.pyplot as plt
import pandas as pd
import openpyxl
import re
import requests
from scipy import stats
import seaborn as sns
#plt.rcParams["figure.figsize"] = (12,10)
#plt.rcParams.update({'font.size': 20})

class covid19italy:
    """"""
    def covid_deaths():
        """Fetch Covid-19 deaths for Italy."""
        # download
        url = 'https://www.epicentro.iss.it/coronavirus/open-data/covid_19-iss.xlsx'
        res = requests.get(url)
        # load
        workbook = openpyxl.load_workbook(io.BytesIO(res.content))
        sheet = workbook['sesso_eta']
        data = {'sex': [], 'age': [], 'deaths': []}
        for value in sheet.iter_rows(min_row=2, max_row=11,
                                     min_col=2, max_col=4, values_only=True):
            sex,age,deaths = value
            data['sex'].append(sex)
            data['age'].append(age)
            data['deaths'].append(deaths)
        for value in sheet.iter_rows(min_row=13, max_row=22,
                                     min_col=2, max_col=4, values_only=True):
            sex,age,deaths = value
            data['sex'].append(sex)
            data['age'].append(age)
            data['deaths'].append(deaths)
        data = pd.DataFrame(data)
        # parse
        data['deaths'] = data.deaths.apply(lambda i: 5 if i == '<5' else int(i))
        def get_age_start(i):
            x = re.match(r'>(90)|(\d+)-\d+', i)
            return int(x[1]) if x[1] is not None else int(x[2])
        data['age_start'] = data.age.apply(get_age_start)
        def get_age_end(i):
            x = re.match(r'>(90)|\d+-(\d+)', i)
            return 99 if x[1] is not None else int(x[2])
        data['age_end'] = data.age.apply(get_age_end)
        # upsample
        cases = {'sex': [], 'age': []}
        for row in data.itertuples():
            random_deaths = stats.multinomial.rvs(row.deaths, [0.1]*10)#, random_state = 12345)
            ages = list(range(row.age_start, row.age_end + 1))
            for age,deaths in zip(ages, random_deaths):
                for _ in range(deaths):
                    cases['sex'].append(row.sex)
                    cases['age'].append(age)
        cases = pd.DataFrame(cases)
        cases['date'] = None
        return cases

def get_data():
    """Get total Covid-19 data for CZ+IT+SE+PL."""
    def _region_to_country(x):
        """Aggregate region data to country data."""
        # aggregate
        return x\
            .groupby(['date','age','sex'])\
            .size()\
            .reset_index(name = 'deaths')
    # Poland
    pl = covid19poland.covid_death_cases(from_github = True)
    pl = _region_to_country(pl)
    pl['country'] = 'PL'
    # Czechia
    cz = covid19czechia.covid_deaths(level = 3, usecache = True)
    cz = _region_to_country(cz)
    cz['country'] = 'CZ'
    # Italy
    it = covid19italy.covid_deaths()
    it['country'] = 'IT'
    # match same dates
    dt_range = pd.date_range(
        start = min(cz.date.min(), pl.date.min()),
        end = min(cz.date.max(), pl.date.max())
    )
    cz = cz[cz.date.isin(dt_range)]
    pl = pl[pl.date.isin(dt_range)]
    # merge
    df = pd.concat([pl, cz])
    # reindex to people
    df = pd.DataFrame(df.values.repeat(df.deaths.apply(int), axis=0), columns=df.columns)
    # merge
    df = pd.concat([df, it])
    # drop deaths column
    df = df.drop(['deaths'], axis = 1)
    df['age'] = df.age.apply(float)
    # return
    return df

def plot_violin(save = False, name="img/parameters/covid_lethality.png"):
    """Violinplot of Covid-19 deaths data per gender and age.
    
    Args:
        save (bool, optional): Whether to save the figure, defaultly not.
        name (str, optional): Path to save the plot to.
    """
    # get data
    df = get_data()
    # plot
    fig1, ax1 = plt.subplots()
    sns.violinplot(x="country", y="age", hue="sex", data = df, ax = ax1)
    # save
    if save: plt.savefig(name)

def test_over60():
    """Test that Covid-19 deaths are significantly greater for > 60 years."""
    # get data
    df = get_data()
    # ttest
    def _ttest_greater(sample_data):
        t,p = stats.ttest_1samp(sample_data, 60)
        return {'t': t, 'p': p/2, 'test': (p/2 < .05 and t > 0)}
    result = {'total': _ttest_greater(df.age), 'per_country': {}, 'per_sex': {}}
    for country,df_country in df.groupby('country'):
        result['per_country'][country] = {'total': _ttest_greater(df_country.age)}
        for sex,df_country_sex in df_country.groupby('sex'):
            result['per_country'][country][sex] = _ttest_greater(df_country_sex.age)
    for sex,df_sex in df.groupby('sex'):
        result['per_sex'][sex] = _ttest_greater(df_sex.age)
    return result
